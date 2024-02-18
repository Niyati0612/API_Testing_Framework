import openpyxl
import pytest
import pytest


@pytest.fixture
def readData():
    global id
    list = []
    path = "../resources/convert1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name("Sheet1")
    row = sheet.max_row
    cols = sheet.max_column

    for r in range(2, row + 1):
        id = sheet.cell(r, 1).value
        name = sheet.cell(r, 3).value

        tuple = (id,name)
        list.append(tuple)
    return list

